from __future__ import annotations

import base64
import hashlib
import hmac
import json
import os
import re
import secrets
import sqlite3
import time
import urllib.error
import urllib.request
import uuid
from io import BytesIO
from datetime import datetime, timedelta, timezone
from functools import wraps
from pathlib import Path
from typing import Any

from flask import Flask, current_app, g, jsonify, request, send_file, send_from_directory
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.styles import Font
from pypinyin import Style, lazy_pinyin


SESSION_COOKIE = "growth_student_session"
PARENT_SESSION_COOKIE = "growth_parent_session"
USERNAME_RE = re.compile(r"^[A-Za-z0-9._-]{2,32}$")
ALLOWED_MEDIA = {
    "image/jpeg": ".jpg",
    "image/png": ".png",
    "image/webp": ".webp",
    "image/gif": ".gif",
    "video/mp4": ".mp4",
    "video/webm": ".webm",
    "video/quicktime": ".mov",
}
LOGIN_ATTEMPTS: dict[str, list[float]] = {}
ADMIN_CACHE: dict[str, tuple[float, dict[str, Any]]] = {}


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def iso_now() -> str:
    return utc_now().isoformat()


def credentials_key() -> bytes:
    secret = str(
        current_app.config.get("SETTINGS_ENCRYPTION_SECRET")
        or current_app.config.get("ADMIN_TEST_TOKEN")
        or current_app.config["DATABASE"]
    )
    return hashlib.sha256(("student-credentials:" + secret).encode("utf-8")).digest()


def seal_credential(value: str) -> str:
    key = credentials_key()
    nonce = os.urandom(16)
    raw = value.encode("utf-8")
    stream = bytearray()
    counter = 0
    while len(stream) < len(raw):
        stream.extend(hmac.new(key, nonce + counter.to_bytes(4, "big"), hashlib.sha256).digest())
        counter += 1
    cipher = bytes(left ^ right for left, right in zip(raw, stream))
    tag = hmac.new(key, nonce + cipher, hashlib.sha256).digest()
    return "enc:v1:" + base64.urlsafe_b64encode(nonce + cipher + tag).decode("ascii")


def open_credential(value: str | None) -> str | None:
    if not value:
        return None
    if not value.startswith("enc:v1:"):
        return value
    try:
        payload = base64.urlsafe_b64decode(value.removeprefix("enc:v1:").encode("ascii"))
        nonce, cipher, tag = payload[:16], payload[16:-32], payload[-32:]
        key = credentials_key()
        expected = hmac.new(key, nonce + cipher, hashlib.sha256).digest()
        if not hmac.compare_digest(tag, expected):
            return None
        stream = bytearray()
        counter = 0
        while len(stream) < len(cipher):
            stream.extend(hmac.new(key, nonce + counter.to_bytes(4, "big"), hashlib.sha256).digest())
            counter += 1
        return bytes(left ^ right for left, right in zip(cipher, stream)).decode("utf-8")
    except (ValueError, TypeError):
        return None


def create_app(test_config: dict[str, Any] | None = None) -> Flask:
    app = Flask(__name__)
    app.config.from_mapping(
        DATABASE=os.getenv("GROWTH_DATABASE", "/var/lib/growth-journal/growth.db"),
        UPLOAD_ROOT=os.getenv("GROWTH_UPLOAD_ROOT", "/var/lib/growth-journal/uploads"),
        SUPABASE_URL=os.getenv("SUPABASE_URL", ""),
        SUPABASE_PUBLISHABLE_KEY=os.getenv("SUPABASE_PUBLISHABLE_KEY", ""),
        DEFAULT_CLASS_ID=os.getenv("DEFAULT_CLASS_ID", "python-summer"),
        ADMIN_TEST_TOKEN=os.getenv("ADMIN_TEST_TOKEN", ""),
        SETTINGS_ENCRYPTION_SECRET=os.getenv("GROWTH_SETTINGS_SECRET", ""),
        MAX_CONTENT_LENGTH=220 * 1024 * 1024,
        SESSION_DAYS=7,
    )
    if test_config:
        app.config.update(test_config)
    app.config["DATABASE"] = str(Path(app.config["DATABASE"]).resolve())
    app.config["UPLOAD_ROOT"] = str(Path(app.config["UPLOAD_ROOT"]).resolve())
    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)  # type: ignore[method-assign]

    @app.teardown_appcontext
    def close_db(_error=None):
        connection = g.pop("db", None)
        if connection is not None:
            connection.close()

    Path(app.config["DATABASE"]).parent.mkdir(parents=True, exist_ok=True)
    Path(app.config["UPLOAD_ROOT"]).mkdir(parents=True, exist_ok=True)
    with app.app_context():
        init_db()
        migrate_demo_accounts()

    register_routes(app)
    from portal_features import register_portal_features
    register_portal_features(
        app,
        get_db=get_db,
        require_admin=require_admin,
        require_student=require_student,
        require_parent=require_parent,
        account_json=account_json,
        student_from_cookie=student_from_cookie,
        verify_admin_token=verify_admin_token,
        json_body=json_body,
        iso_now=iso_now,
    )
    return app


def get_db() -> sqlite3.Connection:
    if "db" not in g:
        connection = sqlite3.connect(current_app.config["DATABASE"], timeout=20)
        connection.row_factory = sqlite3.Row
        connection.execute("pragma foreign_keys = on")
        connection.execute("pragma journal_mode = delete" if current_app.config.get("TESTING") else "pragma journal_mode = wal")
        g.db = connection
    return g.db


def init_db() -> None:
    db = get_db()
    db.executescript(
        """
        create table if not exists student_accounts (
          id text primary key,
          student_id text,
          student_name text not null,
          username text not null collate nocase unique,
          password_hash text not null,
          class_ids text not null default '[]',
          active integer not null default 1 check (active in (0, 1)),
          created_at text not null,
          updated_at text not null
        );
        create table if not exists student_sessions (
          token_hash text primary key,
          account_id text not null references student_accounts(id) on delete cascade,
          created_at text not null,
          expires_at text not null,
          last_seen_at text not null
        );
        create table if not exists parent_sessions (
          token_hash text primary key,
          account_id text not null references student_accounts(id) on delete cascade,
          created_at text not null,
          expires_at text not null,
          last_seen_at text not null
        );
        create table if not exists media_uploads (
          id text primary key,
          owner_kind text not null check (owner_kind in ('admin', 'student')),
          owner_id text not null,
          relative_path text not null unique,
          original_name text not null,
          mime_type text not null,
          size_bytes integer not null,
          created_at text not null
        );
        create index if not exists student_accounts_student_id_idx on student_accounts(student_id);
        create index if not exists student_sessions_account_id_idx on student_sessions(account_id);
        create index if not exists student_sessions_expires_at_idx on student_sessions(expires_at);
        create index if not exists parent_sessions_account_id_idx on parent_sessions(account_id);
        create index if not exists parent_sessions_expires_at_idx on parent_sessions(expires_at);
        """
    )
    columns = {row["name"] for row in db.execute("pragma table_info(student_accounts)").fetchall()}
    if "password_cipher" not in columns:
        db.execute("alter table student_accounts add column password_cipher text")
    if "credentials_assigned" not in columns:
        db.execute("alter table student_accounts add column credentials_assigned integer not null default 1")
    db.commit()


def migrate_demo_accounts() -> None:
    db = get_db()
    class_id = current_app.config["DEFAULT_CLASS_ID"]
    seeds = [
        ("demo-student-01", "林小满", "student01", "scrypt:32768:8:1$e9e4eb222d559986$dc7dca2f1cecb5f0738d09cfd063f70ac5b5b415e85a0746c539d4238b23f22b071fa65127bb2ae7014899b17f71f8c163db8248eab337fb3ada26b7d27c24db"),
        ("demo-student-02", "陈星野", "student02", "scrypt:32768:8:1$9f50054b0971492f$f3b6b90b068269c331939dd87d932e87355c7761f55ec472540524b88eeede15eeaaf6416b5e46ebd4cec3fe459bf4b96bdad2fd5f17e942ffcb9da7f4f00de0"),
    ]
    for account_id, name, username, password_hash in seeds:
        exists = db.execute("select 1 from student_accounts where username = ?", (username,)).fetchone()
        if exists:
            db.execute(
                "update student_accounts set password_cipher=coalesce(password_cipher, ?), credentials_assigned=1 where username=?",
                (seal_credential("123456"), username),
            )
            continue
        now = iso_now()
        db.execute(
            """insert into student_accounts
               (id, student_id, student_name, username, password_hash, password_cipher, credentials_assigned,
                class_ids, active, created_at, updated_at)
               values (?, null, ?, ?, ?, ?, 1, ?, 1, ?, ?)""",
            (account_id, name, username, password_hash, seal_credential("123456"), json.dumps([class_id]), now, now),
        )
    db.commit()


def account_json(row: sqlite3.Row, include_secret: bool = False) -> dict[str, Any]:
    keys = set(row.keys())
    assigned = bool(row["credentials_assigned"]) if "credentials_assigned" in keys else True
    result = {
        "id": row["id"],
        "studentId": row["student_id"],
        "studentName": row["student_name"],
        "username": row["username"],
        "classIds": json.loads(row["class_ids"] or "[]"),
        "active": bool(row["active"]),
        "credentialsAssigned": assigned,
        "parentUsername": f"a{row['username']}" if assigned else None,
        "parentName": f"{row['student_name']}家长",
        "points": int(row["points"] or 0) if "points" in keys else 0,
        "deletedAt": row["deleted_at"] if "deleted_at" in keys else None,
        "createdAt": row["created_at"],
        "updatedAt": row["updated_at"],
    }
    if include_secret:
        result["currentPassword"] = open_credential(row["password_cipher"]) if "password_cipher" in keys and assigned else None
    return result


def json_body() -> dict[str, Any]:
    value = request.get_json(silent=True)
    return value if isinstance(value, dict) else {}


def token_hash(token: str) -> str:
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def student_from_cookie() -> sqlite3.Row | None:
    token = request.cookies.get(SESSION_COOKIE, "")
    if not token:
        return None
    now = iso_now()
    db = get_db()
    row = db.execute(
        """select a.* from student_sessions s
           join student_accounts a on a.id = s.account_id
           where s.token_hash = ? and s.expires_at > ? and a.active = 1""",
        (token_hash(token), now),
    ).fetchone()
    if row:
        db.execute("update student_sessions set last_seen_at = ? where token_hash = ?", (now, token_hash(token)))
        db.commit()
    return row


def require_student(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        account = student_from_cookie()
        if not account:
            return jsonify({"error": "请先登录学生账号。"}), 401
        g.student_account = account
        return view(*args, **kwargs)

    return wrapped


def parent_from_cookie() -> sqlite3.Row | None:
    token = request.cookies.get(PARENT_SESSION_COOKIE, "")
    if not token:
        return None
    now = iso_now()
    db = get_db()
    row = db.execute(
        """select a.* from parent_sessions s
           join student_accounts a on a.id = s.account_id
           where s.token_hash = ? and s.expires_at > ? and a.active = 1
             and a.credentials_assigned = 1 and a.deleted_at is null
             and a.student_id is not null and trim(a.student_id) <> ''""",
        (token_hash(token), now),
    ).fetchone()
    if row:
        db.execute("update parent_sessions set last_seen_at = ? where token_hash = ?", (now, token_hash(token)))
        db.commit()
    return row


def require_parent(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        account = parent_from_cookie()
        if not account:
            return jsonify({"error": "请先登录家长账号。"}), 401
        g.parent_student_account = account
        return view(*args, **kwargs)

    return wrapped


def verify_admin_token() -> dict[str, Any] | None:
    header = request.headers.get("Authorization", "")
    if not header.startswith("Bearer "):
        return None
    token = header[7:].strip()
    if not token:
        return None
    test_token = current_app.config.get("ADMIN_TEST_TOKEN")
    if test_token and secrets.compare_digest(token, test_token):
        return {"id": "test-admin", "email": "admin@test.local"}
    cached = ADMIN_CACHE.get(token_hash(token))
    if cached and cached[0] > time.time():
        return cached[1]
    url = current_app.config["SUPABASE_URL"].rstrip("/") + "/auth/v1/user"
    key = current_app.config["SUPABASE_PUBLISHABLE_KEY"]
    if not url.startswith("http") or not key:
        return None
    outbound = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}", "apikey": key})
    try:
        with urllib.request.urlopen(outbound, timeout=8) as response:
            user = json.loads(response.read().decode("utf-8"))
    except (urllib.error.URLError, TimeoutError, ValueError):
        return None
    if not user.get("id"):
        return None
    ADMIN_CACHE[token_hash(token)] = (time.time() + 300, user)
    return user


def require_admin(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        user = verify_admin_token()
        if not user:
            return jsonify({"error": "管理员登录已失效，请重新登录。"}), 401
        g.admin_user = user
        return view(*args, **kwargs)

    return wrapped


def valid_password(value: Any) -> bool:
    return isinstance(value, str) and 6 <= len(value) <= 128


def valid_username(value: Any) -> bool:
    return isinstance(value, str) and bool(USERNAME_RE.fullmatch(value.strip()))


def normalized_class_ids(value: Any) -> list[str]:
    if not isinstance(value, list):
        return []
    return list(dict.fromkeys(str(item).strip() for item in value if str(item).strip()))[:20]


def login_is_limited(key: str) -> bool:
    cutoff = time.time() - 900
    attempts = [stamp for stamp in LOGIN_ATTEMPTS.get(key, []) if stamp > cutoff]
    LOGIN_ATTEMPTS[key] = attempts
    return len(attempts) >= 8


def record_login_failure(key: str) -> None:
    LOGIN_ATTEMPTS.setdefault(key, []).append(time.time())


def update_supabase_student(student_id: str, name: str) -> None:
    header = request.headers.get("Authorization", "")
    url = current_app.config["SUPABASE_URL"].rstrip("/") + f"/rest/v1/students?id=eq.{student_id}"
    key = current_app.config["SUPABASE_PUBLISHABLE_KEY"]
    if not url.startswith("http") or not key or not header:
        return
    outbound = urllib.request.Request(
        url,
        data=json.dumps({"name": name}).encode("utf-8"),
        method="PATCH",
        headers={"Authorization": header, "apikey": key, "Content-Type": "application/json", "Prefer": "return=minimal"},
    )
    try:
        urllib.request.urlopen(outbound, timeout=8).close()
    except urllib.error.URLError:
        pass


def register_routes(app: Flask) -> None:
    @app.errorhandler(413)
    def too_large(_error):
        return jsonify({"error": "文件过大，视频不能超过 200MB。"}), 413

    def ensure_points_schema() -> None:
        db = get_db()
        columns = {row["name"] for row in db.execute("pragma table_info(student_accounts)").fetchall()}
        if "points" not in columns:
            db.execute("alter table student_accounts add column points integer not null default 0")
        db.execute(
            """create table if not exists point_events (
                id text primary key,
                account_id text not null references student_accounts(id) on delete cascade,
                delta integer not null,
                reason text not null default '',
                source_id text,
                created_at text not null,
                unique(account_id, source_id)
            )"""
        )
        db.commit()

    with app.app_context():
        ensure_points_schema()

    @app.get("/api/health")
    def health():
        return jsonify({"ok": True, "service": "growth-journal-api"})

    @app.get("/uploads/<path:filename>")
    def uploaded_file(filename: str):
        return send_from_directory(current_app.config["UPLOAD_ROOT"], filename, conditional=True)

    @app.post("/api/student/login")
    def student_login():
        body = json_body()
        username = str(body.get("username", "")).strip()
        password = str(body.get("password", ""))
        rate_key = f"{request.remote_addr}:{username.lower()}"
        if login_is_limited(rate_key):
            return jsonify({"error": "尝试次数过多，请 15 分钟后再试。"}), 429
        row = get_db().execute("select * from student_accounts where username = ?", (username,)).fetchone()
        if not row or not row["active"] or not check_password_hash(row["password_hash"], password):
            record_login_failure(rate_key)
            time.sleep(0.25)
            return jsonify({"error": "账号或密码不正确。"}), 401
        LOGIN_ATTEMPTS.pop(rate_key, None)
        token = secrets.token_urlsafe(48)
        now = utc_now()
        expires = now + timedelta(days=int(current_app.config["SESSION_DAYS"]))
        db = get_db()
        db.execute("delete from student_sessions where expires_at <= ?", (now.isoformat(),))
        db.execute(
            "insert into student_sessions (token_hash, account_id, created_at, expires_at, last_seen_at) values (?, ?, ?, ?, ?)",
            (token_hash(token), row["id"], now.isoformat(), expires.isoformat(), now.isoformat()),
        )
        db.commit()
        response = jsonify({"account": account_json(row)})
        response.set_cookie(
            SESSION_COOKIE,
            token,
            max_age=int(current_app.config["SESSION_DAYS"]) * 86400,
            httponly=True,
            secure=request.is_secure,
            samesite="Lax",
            path="/",
        )
        return response

    @app.get("/api/student/session")
    @require_student
    def student_session():
        return jsonify({"account": account_json(g.student_account)})

    @app.post("/api/student/logout")
    def student_logout():
        token = request.cookies.get(SESSION_COOKIE, "")
        if token:
            db = get_db()
            db.execute("delete from student_sessions where token_hash = ?", (token_hash(token),))
            db.commit()
        response = jsonify({"ok": True})
        response.delete_cookie(SESSION_COOKIE, path="/", samesite="Lax")
        return response

    @app.post("/api/parent/login")
    def parent_login():
        body = json_body()
        parent_username = str(body.get("username", "")).strip()
        password = str(body.get("password", ""))
        rate_key = f"parent:{request.remote_addr}:{parent_username.lower()}"
        if login_is_limited(rate_key):
            return jsonify({"error": "尝试次数过多，请 15 分钟后再试。"}), 429
        if not parent_username.lower().startswith("a") or len(parent_username) < 4:
            record_login_failure(rate_key)
            return jsonify({"error": "家长账号或密码不正确。"}), 401
        student_username = parent_username[1:]
        row = get_db().execute("select * from student_accounts where username = ?", (student_username,)).fetchone()
        if (
            not row
            or not row["active"]
            or row["deleted_at"]
            or not row["credentials_assigned"]
            or not str(row["student_id"] or "").strip()
            or not check_password_hash(row["password_hash"], password)
        ):
            record_login_failure(rate_key)
            time.sleep(0.25)
            return jsonify({"error": "家长账号或密码不正确。"}), 401
        LOGIN_ATTEMPTS.pop(rate_key, None)
        token = secrets.token_urlsafe(48)
        now = utc_now()
        expires = now + timedelta(days=int(current_app.config["SESSION_DAYS"]))
        db = get_db()
        db.execute("delete from parent_sessions where expires_at <= ?", (now.isoformat(),))
        db.execute(
            "insert into parent_sessions (token_hash, account_id, created_at, expires_at, last_seen_at) values (?, ?, ?, ?, ?)",
            (token_hash(token), row["id"], now.isoformat(), expires.isoformat(), now.isoformat()),
        )
        db.commit()
        response = jsonify({
            "parent": {"username": f"a{row['username']}", "name": f"{row['student_name']}家长"},
            "student": account_json(row),
        })
        response.set_cookie(
            PARENT_SESSION_COOKIE,
            token,
            max_age=int(current_app.config["SESSION_DAYS"]) * 86400,
            httponly=True,
            secure=request.is_secure,
            samesite="Lax",
            path="/",
        )
        return response

    @app.get("/api/parent/session")
    @require_parent
    def parent_session():
        row = g.parent_student_account
        return jsonify({
            "parent": {"username": f"a{row['username']}", "name": f"{row['student_name']}家长"},
            "student": account_json(row),
        })

    @app.post("/api/parent/logout")
    def parent_logout():
        token = request.cookies.get(PARENT_SESSION_COOKIE, "")
        if token:
            db = get_db()
            db.execute("delete from parent_sessions where token_hash = ?", (token_hash(token),))
            db.commit()
        response = jsonify({"ok": True})
        response.delete_cookie(PARENT_SESSION_COOKIE, path="/", samesite="Lax")
        return response

    @app.post("/api/student/password")
    @require_student
    def student_password():
        body = json_body()
        current = str(body.get("currentPassword", ""))
        next_password = body.get("nextPassword")
        row = g.student_account
        if not check_password_hash(row["password_hash"], current):
            return jsonify({"error": "当前密码不正确。"}), 400
        if not valid_password(next_password):
            return jsonify({"error": "新密码需要 6–128 位。"}), 400
        db = get_db()
        db.execute(
            "update student_accounts set password_hash = ?, password_cipher = ?, updated_at = ? where id = ?",
            (generate_password_hash(next_password), seal_credential(next_password), iso_now(), row["id"]),
        )
        db.execute("delete from student_sessions where account_id = ? and token_hash <> ?", (row["id"], token_hash(request.cookies[SESSION_COOKIE])))
        db.execute("delete from parent_sessions where account_id = ?", (row["id"],))
        db.commit()
        return jsonify({"ok": True})

    @app.get("/api/admin/student-accounts")
    @require_admin
    def list_accounts():
        class_id = request.args.get("class_id", "").strip()
        rows = get_db().execute("select * from student_accounts order by student_name collate nocase, created_at").fetchall()
        accounts = [account_json(row, include_secret=True) for row in rows]
        for item, row in zip(accounts, rows):
            item["points"] = int(row["points"] or 0)
        if class_id:
            accounts = [item for item in accounts if class_id in item["classIds"]]
        return jsonify({"accounts": accounts})

    @app.post("/api/admin/student-accounts")
    @require_admin
    def create_account():
        body = json_body()
        username = str(body.get("username", "")).strip()
        student_name = str(body.get("studentName", "")).strip()
        password = body.get("password")
        class_ids = normalized_class_ids(body.get("classIds"))
        if not student_name or len(student_name) > 80:
            return jsonify({"error": "请填写学生姓名。"}), 400
        if not valid_username(username):
            return jsonify({"error": "账号需为 2–32 位字母、数字、点、横线或下划线。"}), 400
        if not valid_password(password):
            return jsonify({"error": "密码需要 6–128 位。"}), 400
        if not class_ids:
            return jsonify({"error": "请选择学生所属班级。"}), 400
        now = iso_now()
        account_id = str(uuid.uuid4())
        student_id = str(body.get("studentId") or "").strip() or None
        try:
            db = get_db()
            placeholder = db.execute(
                "select * from student_accounts where student_id=? and credentials_assigned=0",
                (student_id,),
            ).fetchone() if student_id else None
            if placeholder:
                account_id = placeholder["id"]
                db.execute(
                    """update student_accounts set student_name=?, username=?, password_hash=?, password_cipher=?,
                       credentials_assigned=1, class_ids=?, active=1, deleted_at=null, updated_at=? where id=?""",
                    (student_name, username, generate_password_hash(password), seal_credential(password),
                     json.dumps(class_ids), now, account_id),
                )
            else:
                db.execute(
                    """insert into student_accounts
                       (id, student_id, student_name, username, password_hash, password_cipher, credentials_assigned,
                        class_ids, active, created_at, updated_at)
                       values (?, ?, ?, ?, ?, ?, 1, ?, 1, ?, ?)""",
                    (account_id, student_id, student_name, username, generate_password_hash(password),
                     seal_credential(password), json.dumps(class_ids), now, now),
                )
            db.commit()
        except sqlite3.IntegrityError:
            return jsonify({"error": "这个登录账号已经存在。"}), 409
        row = get_db().execute("select * from student_accounts where id = ?", (account_id,)).fetchone()
        return jsonify({"account": account_json(row, include_secret=True)}), 201

    def generated_username(name: str, occupied: set[str]) -> str:
        base = "".join(lazy_pinyin(name, style=Style.FIRST_LETTER)).lower()
        base = re.sub(r"[^a-z0-9._-]", "", base)[:28] or "student"
        if len(base) < 2:
            base = f"{base}s"
        candidate = base
        number = 1
        while candidate.lower() in occupied:
            candidate = f"{base[:28]}{number}"
            number += 1
        occupied.add(candidate.lower())
        return candidate

    @app.get("/api/admin/student-import/template")
    @require_admin
    def student_import_template():
        from openpyxl import Workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "学生名单"
        sheet.append(["学生姓名"])
        sheet.column_dimensions["A"].width = 34
        sheet["A1"].font = Font(bold=True)
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name="学生批量导入模板.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    @app.post("/api/admin/student-import/preview")
    @require_admin
    def student_import_preview():
        uploaded = request.files.get("file")
        if not uploaded or not uploaded.filename or not uploaded.filename.lower().endswith(".xlsx"):
            return jsonify({"error": "请选择 .xlsx 学生名单。"}), 400
        try:
            workbook = load_workbook(uploaded.stream, read_only=True, data_only=True)
            sheet = workbook.active
            values = list(sheet.iter_rows(values_only=True))
            workbook.close()
        except Exception:
            return jsonify({"error": "Excel 文件无法读取，请使用系统提供的模板。"}), 400
        header_row = -1
        name_column = -1
        for row_index, row in enumerate(values[:10]):
            for column_index, value in enumerate(row):
                if str(value or "").strip() == "学生姓名":
                    header_row, name_column = row_index, column_index
                    break
            if header_row >= 0:
                break
        if header_row < 0:
            return jsonify({"error": "没有找到“学生姓名”表头，请下载模板后填写。"}), 400
        occupied = {str(row["username"]).lower() for row in get_db().execute("select username from student_accounts").fetchall()}
        rows: list[dict[str, Any]] = []
        blank_count = 0
        for row_index, row in enumerate(values[header_row + 1:], start=header_row + 2):
            raw = row[name_column] if name_column < len(row) else None
            name = str(raw or "").strip()
            if not name:
                blank_count += 1
                continue
            if len(name) > 80:
                return jsonify({"error": f"第 {row_index} 行姓名超过80个字符。"}), 400
            username = generated_username(name, occupied)
            rows.append({"rowNumber": row_index, "studentName": name, "username": username,
                         "parentUsername": f"a{username}", "password": "123456"})
            if len(rows) > 500:
                return jsonify({"error": "一次最多导入500名学生。"}), 400
        if not rows:
            return jsonify({"error": "表格中没有可导入的学生姓名。"}), 400
        return jsonify({"rows": rows, "blankRowsSkipped": blank_count})

    @app.post("/api/admin/student-import/commit")
    @require_admin
    def student_import_commit():
        body = json_body()
        class_ids = normalized_class_ids([body.get("classId")])
        students = body.get("students")
        if not class_ids or not isinstance(students, list) or not students or len(students) > 500:
            return jsonify({"error": "请选择班级并提交1—500名学生。"}), 400
        prepared: list[tuple[str, str, str]] = []
        usernames: set[str] = set()
        for item in students:
            if not isinstance(item, dict):
                return jsonify({"error": "导入数据格式不正确。"}), 400
            student_id = str(item.get("studentId") or "").strip()
            student_name = str(item.get("studentName") or "").strip()
            username = str(item.get("username") or "").strip()
            if not student_id or not student_name or len(student_name) > 80 or not valid_username(username):
                return jsonify({"error": f"学生“{student_name or '未知'}”的数据不完整。"}), 400
            if username.lower() in usernames:
                return jsonify({"error": f"账号 {username} 重复。"}), 409
            usernames.add(username.lower())
            prepared.append((student_id, student_name, username))
        db = get_db()
        placeholders = ",".join("?" for _ in usernames)
        existing = db.execute(f"select username from student_accounts where lower(username) in ({placeholders})", tuple(usernames)).fetchone()
        if existing:
            return jsonify({"error": f"账号 {existing['username']} 已存在，请重新预览名单。"}), 409
        now = iso_now()
        created_ids: list[str] = []
        try:
            for student_id, student_name, username in prepared:
                account_id = str(uuid.uuid4())
                db.execute(
                    """insert into student_accounts
                       (id,student_id,student_name,username,password_hash,password_cipher,credentials_assigned,
                        class_ids,active,created_at,updated_at)
                       values (?,?,?,?,?,?,1,?,1,?,?)""",
                    (account_id, student_id, student_name, username, generate_password_hash("123456"),
                     seal_credential("123456"), json.dumps(class_ids), now, now),
                )
                created_ids.append(account_id)
            db.commit()
        except sqlite3.IntegrityError:
            db.rollback()
            return jsonify({"error": "导入过程中出现账号冲突，请重新预览后再试。"}), 409
        rows = [db.execute("select * from student_accounts where id=?", (account_id,)).fetchone() for account_id in created_ids]
        return jsonify({"accounts": [account_json(row, include_secret=True) for row in rows]}), 201

    @app.patch("/api/admin/student-accounts/<account_id>")
    @require_admin
    def update_account(account_id: str):
        db = get_db()
        row = db.execute("select * from student_accounts where id = ?", (account_id,)).fetchone()
        if not row:
            return jsonify({"error": "没有找到这个学生账号。"}), 404
        body = json_body()
        student_name = str(body.get("studentName", row["student_name"])).strip()
        username = str(body.get("username", row["username"])).strip()
        student_id = body.get("studentId", row["student_id"])
        student_id = str(student_id).strip() if student_id else None
        class_ids = normalized_class_ids(body.get("classIds", json.loads(row["class_ids"])))
        active = 1 if bool(body.get("active", bool(row["active"]))) else 0
        if not student_name or len(student_name) > 80 or not valid_username(username) or not class_ids:
            return jsonify({"error": "请检查姓名、登录账号和班级。"}), 400
        try:
            db.execute(
                """update student_accounts set student_id = ?, student_name = ?, username = ?, class_ids = ?, active = ?, updated_at = ?
                   where id = ?""",
                (student_id, student_name, username, json.dumps(class_ids), active, iso_now(), account_id),
            )
            if not active:
                db.execute("delete from student_sessions where account_id = ?", (account_id,))
            db.commit()
        except sqlite3.IntegrityError:
            return jsonify({"error": "这个登录账号已经存在。"}), 409
        if student_id and student_name != row["student_name"]:
            update_supabase_student(student_id, student_name)
        updated = db.execute("select * from student_accounts where id = ?", (account_id,)).fetchone()
        return jsonify({"account": account_json(updated, include_secret=True)})

    @app.post("/api/admin/student-accounts/<account_id>/reset-password")
    @require_admin
    def reset_account_password(account_id: str):
        password = json_body().get("password")
        if not valid_password(password):
            return jsonify({"error": "密码需要 6–128 位。"}), 400
        db = get_db()
        result = db.execute(
            """update student_accounts set password_hash = ?, password_cipher = ?, credentials_assigned=1,
               active=1, updated_at = ? where id = ?""",
            (generate_password_hash(password), seal_credential(password), iso_now(), account_id),
        )
        db.execute("delete from student_sessions where account_id = ?", (account_id,))
        db.execute("delete from parent_sessions where account_id = ?", (account_id,))
        db.commit()
        if not result.rowcount:
            return jsonify({"error": "没有找到这个学生账号。"}), 404
        return jsonify({"ok": True})

    @app.post("/api/admin/student-profiles/<student_id>/recycle")
    @require_admin
    def recycle_student_profile(student_id: str):
        body = json_body()
        student_name = str(body.get("studentName", "")).strip()
        class_id = str(body.get("classId", "")).strip()
        if not student_name or not class_id:
            return jsonify({"error": "学生姓名和班级不能为空。"}), 400
        db = get_db()
        row = db.execute("select * from student_accounts where student_id=?", (student_id,)).fetchone()
        now = iso_now()
        if row:
            db.execute(
                "update student_accounts set deleted_at=?, active=0, updated_at=? where id=?",
                (now, now, row["id"]),
            )
            account_id = row["id"]
        else:
            account_id = str(uuid.uuid4())
            placeholder_username = f"archived-{account_id[:20]}"
            placeholder_password = secrets.token_urlsafe(24)
            db.execute(
                """insert into student_accounts
                   (id,student_id,student_name,username,password_hash,password_cipher,credentials_assigned,
                    class_ids,active,created_at,updated_at,deleted_at)
                   values (?,?,?,?,?,null,0,?,0,?,?,?)""",
                (account_id, student_id, student_name, placeholder_username,
                 generate_password_hash(placeholder_password), json.dumps([class_id]), now, now, now),
            )
        db.execute("delete from student_sessions where account_id=?", (account_id,))
        db.execute("delete from parent_sessions where account_id=?", (account_id,))
        db.commit()
        updated = db.execute("select * from student_accounts where id=?", (account_id,)).fetchone()
        return jsonify({"account": account_json(updated, include_secret=True)})

    @app.get("/api/leaderboard")
    def leaderboard():
        student = student_from_cookie()
        if not student:
            return jsonify({"error": "学生登录已失效，请重新登录。"}), 401
        rows = get_db().execute(
            "select id, student_name, points from student_accounts where active = 1 order by points desc, student_name collate nocase"
        ).fetchall()
        return jsonify({"accounts": [
            {"id": row["id"], "studentName": row["student_name"], "points": int(row["points"] or 0)}
            for row in rows
        ]})

    @app.post("/api/student/points/award")
    def student_points_award():
        student = student_from_cookie()
        if not student:
            return jsonify({"error": "学生登录已失效，请重新登录。"}), 401
        body = json_body()
        question_id = str(body.get("questionId", "")).strip()[:160]
        points = 100
        reason = str(body.get("reason", "课件答题")).strip()[:120]
        if not question_id or points <= 0:
            return jsonify({"error": "积分请求无效。"}), 400
        db = get_db()
        awarded = False
        try:
            db.execute(
                "insert into point_events (id, account_id, delta, reason, source_id, created_at) values (?, ?, ?, ?, ?, ?)",
                (str(uuid.uuid4()), student["id"], points, reason, question_id, iso_now()),
            )
            db.execute("update student_accounts set points = points + ?, updated_at = ? where id = ?", (points, iso_now(), student["id"]))
            db.commit()
            awarded = True
        except sqlite3.IntegrityError:
            db.rollback()
        row = db.execute("select points from student_accounts where id = ?", (student["id"],)).fetchone()
        return jsonify({"points": int(row["points"] or 0), "awarded": awarded})

    @app.patch("/api/admin/student-accounts/<account_id>/points")
    @require_admin
    def adjust_account_points(account_id: str):
        body = json_body()
        try:
            delta = int(body.get("delta") or 0)
        except (TypeError, ValueError):
            return jsonify({"error": "积分必须是整数。"}), 400
        if delta == 0 or abs(delta) > 1000:
            return jsonify({"error": "单次积分调整范围为 -1000 到 1000。"}), 400
        reason = str(body.get("reason", "教师调整")).strip()[:120]
        db = get_db()
        row = db.execute("select points from student_accounts where id = ?", (account_id,)).fetchone()
        if not row:
            return jsonify({"error": "学生账号不存在。"}), 404
        old_points = int(row["points"] or 0)
        new_points = max(0, old_points + delta)
        actual_delta = new_points - old_points
        db.execute("update student_accounts set points = ?, updated_at = ? where id = ?", (new_points, iso_now(), account_id))
        db.execute(
            "insert into point_events (id, account_id, delta, reason, source_id, created_at) values (?, ?, ?, ?, ?, ?)",
            (str(uuid.uuid4()), account_id, actual_delta, reason, None, iso_now()),
        )
        db.commit()
        return jsonify({"ok": True, "points": new_points, "delta": actual_delta})

    @app.get("/api/admin/point-events")
    @require_admin
    def list_point_events():
        rows = get_db().execute(
            """select e.id, e.account_id, a.student_name, e.delta, e.reason, e.created_at
               from point_events e join student_accounts a on a.id = e.account_id
               order by e.created_at desc limit 200"""
        ).fetchall()
        return jsonify({"events": [dict(row) for row in rows]})

    @app.post("/api/media/upload")
    def upload_media():
        admin = verify_admin_token()
        student = None if admin else student_from_cookie()
        if not admin and not student:
            return jsonify({"error": "请先登录后再上传。"}), 401
        uploaded = request.files.get("file")
        if not uploaded or not uploaded.filename:
            return jsonify({"error": "请选择要上传的文件。"}), 400
        mime = (uploaded.mimetype or "").lower()
        extension = ALLOWED_MEDIA.get(mime)
        if not extension:
            return jsonify({"error": "只支持常用图片和 MP4、WebM、MOV 视频。"}), 400
        folder = str(request.form.get("folder", "media"))
        folder_parts = [secure_filename(part) for part in folder.split("/") if secure_filename(part)][:5]
        date_parts = utc_now().strftime("%Y/%m").split("/")
        relative_dir = Path(*folder_parts, *date_parts)
        upload_id = str(uuid.uuid4())
        relative_path = (relative_dir / f"{upload_id}{extension}").as_posix()
        root = Path(current_app.config["UPLOAD_ROOT"]).resolve()
        destination = (root / relative_path).resolve()
        if root not in destination.parents:
            return jsonify({"error": "上传目录无效。"}), 400
        destination.parent.mkdir(parents=True, exist_ok=True)
        uploaded.save(destination)
        size = destination.stat().st_size
        max_size = 200 * 1024 * 1024 if mime.startswith("video/") else 25 * 1024 * 1024
        if size <= 0 or size > max_size:
            destination.unlink(missing_ok=True)
            return jsonify({"error": "图片不能超过 25MB，视频不能超过 200MB。"}), 413
        owner_kind = "admin" if admin else "student"
        owner_id = str((admin or {}).get("id") if admin else student["id"])
        db = get_db()
        db.execute(
            """insert into media_uploads
               (id, owner_kind, owner_id, relative_path, original_name, mime_type, size_bytes, created_at)
               values (?, ?, ?, ?, ?, ?, ?, ?)""",
            (upload_id, owner_kind, owner_id, relative_path, uploaded.filename[:240], mime, size, iso_now()),
        )
        db.commit()
        return jsonify({"id": upload_id, "url": f"/uploads/{relative_path}", "path": f"server:{upload_id}", "size": size})

    @app.delete("/api/media/<upload_id>")
    @require_admin
    def delete_media(upload_id: str):
        db = get_db()
        row = db.execute("select * from media_uploads where id = ?", (upload_id,)).fetchone()
        if not row:
            return jsonify({"ok": True})
        root = Path(current_app.config["UPLOAD_ROOT"]).resolve()
        destination = (root / row["relative_path"]).resolve()
        if root in destination.parents:
            destination.unlink(missing_ok=True)
        db.execute("delete from media_uploads where id = ?", (upload_id,))
        db.commit()
        return jsonify({"ok": True})
