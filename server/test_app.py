import io
import json
import sqlite3
import tempfile
import unittest
import zipfile
from pathlib import Path

from PIL import Image

from app import compress_existing_uploaded_images, create_app, get_db
from core_data import migrate_embedded_data_images


class ApiTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        root = Path(self.temp.name)
        self.db_path = root / "test.db"
        self.upload_root = root / "uploads"
        self.app = create_app({
            "TESTING": True,
            "DATABASE": str(self.db_path),
            "UPLOAD_ROOT": str(self.upload_root),
            "ADMIN_TEST_TOKEN": "teacher-test-token",
            "ADMIN_BOOTSTRAP_EMAIL": "teacher@example.com",
            "ADMIN_BOOTSTRAP_PASSWORD": "123456",
            "DEFAULT_CLASS_ID": "class-one",
        })
        self.client = self.app.test_client()

    def tearDown(self):
        self.client._context_stack.close()
        self.temp.cleanup()

    def admin_headers(self):
        return {"Authorization": "Bearer teacher-test-token"}

    def test_self_hosted_admin_login_and_core_data_crud(self):
        login = self.client.post("/api/admin/login", json={"email": "teacher@example.com", "password": "123456"})
        self.assertEqual(login.status_code, 200, login.json)
        self.assertIn("HttpOnly", login.headers["Set-Cookie"])
        self.assertEqual(self.client.get("/api/admin/session").status_code, 200)

        class_response = self.client.post("/api/admin/classes", json={"name": "Python班", "description": "测试"})
        self.assertEqual(class_response.status_code, 201, class_response.json)
        class_id = class_response.json["item"]["id"]
        term_response = self.client.post("/api/admin/terms", json={
            "class_id": class_id, "name": "2026秋季", "start_date": "2026-09-01", "end_date": "2027-01-31",
        })
        self.assertEqual(term_response.status_code, 201, term_response.json)
        term_id = term_response.json["item"]["id"]
        student_response = self.client.post("/api/admin/students", json={
            "class_id": class_id, "name": "测试学生", "joined_on": "2026-09-03",
        })
        self.assertEqual(student_response.status_code, 201, student_response.json)
        student_id = student_response.json["item"]["id"]
        lesson_response = self.client.post("/api/admin/lessons", json={
            "class_id": class_id, "term_id": term_id, "sequence_no": 1, "title": "第一课",
            "lesson_date": "2026-09-03", "summary": "测试", "content_html": "<p>内容</p>",
        })
        self.assertEqual(lesson_response.status_code, 201, lesson_response.json)
        lesson_id = lesson_response.json["item"]["id"]
        lesson_list = self.client.get(f"/api/data/lessons?term_id={term_id}")
        self.assertEqual(lesson_list.json["items"][0]["content_html"], "")
        lesson_detail = self.client.get(f"/api/data/lessons/{lesson_id}")
        self.assertEqual(lesson_detail.json["item"]["content_html"], "<p>内容</p>")
        record_response = self.client.put("/api/admin/records", json={
            "lesson_id": lesson_id, "student_id": student_id, "comment": "很好",
            "thinking_score": 5, "focus_score": 4, "creativity_score": 3,
            "coding_score": 5, "motivation_score": 4,
        })
        self.assertEqual(record_response.status_code, 200, record_response.json)
        listed = self.client.get(f"/api/data/students?class_id={class_id}")
        self.assertEqual(listed.status_code, 200)
        self.assertEqual(listed.json["items"][0]["name"], "测试学生")
        self.assertEqual(self.client.post("/api/admin/logout").status_code, 200)
        self.assertEqual(self.client.get("/api/admin/session").status_code, 401)

    def test_embedded_data_image_is_extracted_and_compressed(self):
        pixel = (
            "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAusB9Y9Zl2sAAAAASUVORK5CYII="
        )
        with self.app.app_context():
            db = get_db()
            db.execute("insert into classes values ('image-class','图片班',null,'2026-09-03')")
            db.execute("insert into terms values ('image-term','image-class','秋季','2026-09-01','2027-01-01','2026-09-03')")
            db.execute(
                """insert into lessons
                   (id,class_id,term_id,sequence_no,title,lesson_date,summary,content_html,created_at)
                   values ('image-lesson','image-class','image-term',1,'图片课','2026-09-03',null,?,'2026-09-03')""",
                (f'<p>图片</p><img src="data:image/png;base64,{pixel}">',),
            )
            db.commit()
            self.assertEqual(migrate_embedded_data_images(db, self.upload_root), 1)
            html = db.execute("select content_html from lessons where id='image-lesson'").fetchone()[0]
            self.assertNotIn("data:image", html)
            self.assertIn("/uploads/migrated-media/", html)
            relative = html.split('/uploads/', 1)[1].split('"', 1)[0]
            self.assertTrue((self.upload_root / relative).is_file())

    def test_migrated_password_is_hashed_and_login_is_shared_server_side(self):
        db = sqlite3.connect(self.db_path)
        try:
            password_hash = db.execute("select password_hash from student_accounts where username='student01'").fetchone()[0]
        finally:
            db.close()
        self.assertNotEqual(password_hash, "123456")
        self.assertTrue(password_hash.startswith("scrypt:"))

        login = self.client.post("/api/student/login", json={"username": "student01", "password": "123456"})
        self.assertEqual(login.status_code, 200)
        self.assertIn("HttpOnly", login.headers["Set-Cookie"])
        self.assertEqual(self.client.get("/api/student/session").status_code, 200)

    def test_teacher_can_create_edit_disable_and_reset_account(self):
        created = self.client.post("/api/admin/student-accounts", headers=self.admin_headers(), json={
            "studentId": "profile-1", "studentName": "张同学", "username": "zhang01",
            "password": "abc12345", "classIds": ["class-one"],
        })
        self.assertEqual(created.status_code, 201)
        account_id = created.json["account"]["id"]
        edited = self.client.patch(f"/api/admin/student-accounts/{account_id}", headers=self.admin_headers(), json={
            "studentName": "张小满", "username": "zhang02", "classIds": ["class-one"], "active": False,
        })
        self.assertEqual(edited.status_code, 200)
        self.assertFalse(edited.json["account"]["active"])
        reset = self.client.post(f"/api/admin/student-accounts/{account_id}/reset-password", headers=self.admin_headers(), json={"password": "newpass88"})
        self.assertEqual(reset.status_code, 200)

    def test_logged_in_student_can_upload_media(self):
        self.client.post("/api/student/login", json={"username": "student01", "password": "123456"})
        original = io.BytesIO()
        Image.new("RGB", (2400, 1600), "#4c8df6").save(original, "PNG")
        original.seek(0)
        response = self.client.post("/api/media/upload", data={
            "folder": "homework/demo",
            "file": (original, "work.png", "image/png"),
        }, content_type="multipart/form-data")
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json["url"].startswith("/uploads/homework/demo/"))
        self.assertTrue(response.json["url"].endswith(".webp"))
        downloaded = self.client.get(response.json["url"])
        self.assertEqual(downloaded.status_code, 200)
        self.assertIn("immutable", downloaded.headers["Cache-Control"])
        with Image.open(io.BytesIO(downloaded.data)) as image:
            self.assertEqual(image.format, "WEBP")
            self.assertLessEqual(max(image.size), 1800)
        downloaded.close()

    def test_admin_can_upload_pdf_attachment(self):
        pdf = io.BytesIO(b"%PDF-1.4\n% test attachment\n")
        response = self.client.post("/api/media/upload", headers=self.admin_headers(), data={
            "folder": "lessons/term-one",
            "file": (pdf, "lesson-notes.pdf", "application/pdf"),
        }, content_type="multipart/form-data")
        self.assertEqual(response.status_code, 200, response.json)
        self.assertTrue(response.json["url"].startswith("/uploads/lessons/term-one/"))
        self.assertTrue(response.json["url"].endswith(".pdf"))
        downloaded = self.client.get(response.json["url"])
        self.assertEqual(downloaded.status_code, 200)
        self.assertEqual(downloaded.data, b"%PDF-1.4\n% test attachment\n")
        downloaded.close()

    def test_admin_can_upload_pdf_to_class_files(self):
        pdf = io.BytesIO(b"%PDF-1.4\n% class file\n")
        response = self.client.post("/api/files", headers=self.admin_headers(), data={
            "classId": "class-one",
            "file": (pdf, "class-notes.pdf", "application/pdf"),
        }, content_type="multipart/form-data")
        self.assertEqual(response.status_code, 200, response.json)
        self.assertEqual(response.json["displayName"], "class-notes.pdf")

    def test_student_can_upload_pdf_with_chinese_filename(self):
        login = self.client.post("/api/student/login", json={"username": "student01", "password": "123456"})
        self.assertEqual(login.status_code, 200)
        pdf = io.BytesIO(b"%PDF-1.4\n% student file\n")
        response = self.client.post("/api/files", data={
            "classId": "class-one",
            "file": (pdf, "第一节课详细讲解笔记.pdf", "application/pdf"),
        }, content_type="multipart/form-data")
        self.assertEqual(response.status_code, 200, response.json)
        self.assertTrue(response.json["displayName"].endswith(".pdf"))

    def test_existing_uploaded_photos_are_compressed_without_changing_url(self):
        relative = "records/example/large.png"
        destination = self.upload_root / relative
        destination.parent.mkdir(parents=True)
        Image.new("RGB", (2400, 1600), "#86d38b").save(destination, "PNG")
        before = destination.stat().st_size
        with self.app.app_context():
            db = get_db()
            db.execute(
                """insert into media_uploads
                   (id,owner_kind,owner_id,relative_path,original_name,mime_type,size_bytes,created_at)
                   values ('existing-image','admin','teacher',?,'large.png','image/png',?,'2026-09-03')""",
                (relative, before),
            )
            db.commit()
            result = compress_existing_uploaded_images(db, self.upload_root)
            self.assertEqual(result["scanned"], 1)
            self.assertEqual(result["optimized"], 1)
            stored_size = db.execute(
                "select size_bytes from media_uploads where id='existing-image'"
            ).fetchone()[0]
        self.assertEqual(stored_size, destination.stat().st_size)
        self.assertLess(destination.stat().st_size, before)
        with Image.open(destination) as image:
            self.assertEqual(image.format, "PNG")
            self.assertLessEqual(max(image.size), 1800)

    def test_parent_uses_prefixed_student_account_and_same_password(self):
        created = self.client.post("/api/admin/student-accounts", headers=self.admin_headers(), json={
            "studentId": "profile-parent", "studentName": "王小星", "username": "wang01",
            "password": "parent88", "classIds": ["class-one"],
        })
        self.assertEqual(created.status_code, 201)
        self.assertEqual(created.json["account"]["parentUsername"], "awang01")
        self.assertEqual(created.json["account"]["parentName"], "王小星家长")
        self.assertEqual(created.json["account"]["currentPassword"], "parent88")
        db = sqlite3.connect(self.db_path)
        try:
            cipher = db.execute("select password_cipher from student_accounts where username='wang01'").fetchone()[0]
        finally:
            db.close()
        self.assertNotEqual(cipher, "parent88")
        login = self.client.post("/api/parent/login", json={"username": "awang01", "password": "parent88"})
        self.assertEqual(login.status_code, 200)
        self.assertEqual(login.json["student"]["studentId"], "profile-parent")
        self.assertEqual(self.client.get("/api/parent/session").status_code, 200)

    def test_parent_login_rejects_account_without_student_profile(self):
        created = self.client.post("/api/admin/student-accounts", headers=self.admin_headers(), json={
            "studentName": "未绑定同学", "username": "unbound01",
            "password": "parent88", "classIds": ["class-one"],
        })
        self.assertEqual(created.status_code, 201)
        login = self.client.post("/api/parent/login", json={"username": "aunbound01", "password": "parent88"})
        self.assertEqual(login.status_code, 401)

    def test_parent_feedback_allows_a_private_conversation_with_teacher(self):
        account = self.client.post("/api/admin/student-accounts", headers=self.admin_headers(), json={
            "studentId": "feedback-profile", "studentName": "反馈同学", "username": "feedback01",
            "password": "parent88", "classIds": ["class-one"],
        })
        self.assertEqual(account.status_code, 201, account.json)
        login = self.client.post("/api/parent/login", json={"username": "afeedback01", "password": "parent88"})
        self.assertEqual(login.status_code, 200, login.json)
        created = self.client.post("/api/parent/feedback", json={"content": "孩子很喜欢课堂，希望多一些循环练习。"})
        self.assertEqual(created.status_code, 201, created.json)
        self.assertEqual(created.json["message"]["author"], "parent")
        message_id = created.json["message"]["id"]
        parent_messages = self.client.get("/api/parent/feedback")
        self.assertEqual(len(parent_messages.json["messages"]), 1)

        teacher_view = self.client.get("/api/admin/parent-feedback?class_id=class-one", headers=self.admin_headers())
        self.assertEqual(teacher_view.status_code, 200, teacher_view.json)
        self.assertEqual(teacher_view.json["messages"][0]["content"], "孩子很喜欢课堂，希望多一些循环练习。")
        reply = self.client.post(
            f"/api/admin/parent-feedback/{message_id}/reply", headers=self.admin_headers(),
            json={"content": "收到，下节课会安排一组循序渐进的挑战题。"},
        )
        self.assertEqual(reply.status_code, 201, reply.json)
        self.assertEqual(reply.json["message"]["author"], "teacher")

        continued = self.client.post("/api/parent/feedback", json={"content": "谢谢老师，我们会继续练习。"})
        self.assertEqual(continued.status_code, 201, continued.json)
        history = self.client.get("/api/parent/feedback").json["messages"]
        self.assertEqual([item["author"] for item in history], ["parent", "teacher", "parent"])

    def test_unbound_profile_can_be_recycled_restored_and_assigned(self):
        recycled = self.client.post("/api/admin/student-profiles/profile-new/recycle", headers=self.admin_headers(), json={
            "studentName": "新同学", "classId": "class-one",
        })
        self.assertEqual(recycled.status_code, 200)
        placeholder = recycled.json["account"]
        self.assertFalse(placeholder["credentialsAssigned"])
        self.assertIsNotNone(placeholder["deletedAt"])
        restored = self.client.post(f"/api/admin/student-accounts/{placeholder['id']}/restore", headers=self.admin_headers())
        self.assertEqual(restored.status_code, 200)
        assigned = self.client.post("/api/admin/student-accounts", headers=self.admin_headers(), json={
            "studentId": "profile-new", "studentName": "新同学", "username": "newstudent",
            "password": "newpass88", "classIds": ["class-one"],
        })
        self.assertEqual(assigned.status_code, 201)
        self.assertEqual(assigned.json["account"]["id"], placeholder["id"])
        self.assertTrue(assigned.json["account"]["credentialsAssigned"])

    def test_class_resource_assignment_is_independent(self):
        state = self.client.get("/api/admin/feature-state?class_id=class-one", headers=self.admin_headers())
        self.assertEqual(state.status_code, 200)
        course = next(item for item in state.json["resourceLibrary"] if item["kind"] == "course")
        assigned = self.client.put(f"/api/admin/class-resources/{course['id']}", headers=self.admin_headers(), json={
            "classId": "class-two", "assigned": True, "enabled": True,
        })
        self.assertEqual(assigned.status_code, 200)
        class_two = self.client.get("/api/admin/feature-state?class_id=class-two", headers=self.admin_headers()).json
        selected = next(item for item in class_two["resourceLibrary"] if item["id"] == course["id"])
        self.assertTrue(selected["assigned"])
        self.assertTrue(selected["enabled"])

    def test_course_zip_upload_syncs_course_and_assessment(self):
        package = io.BytesIO()
        assessment = {
            "title": "第五课课后测评",
            "questions": [{"id": "q1", "type": "choice", "title": "range(5)有几个数？", "options": ["4个", "5个"], "answer": "5个", "points": 10}],
        }
        with zipfile.ZipFile(package, "w", zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("lesson/index.html", "<!doctype html><h1>第五课</h1>")
            archive.writestr("lesson/assessment.json", json.dumps(assessment, ensure_ascii=False))
        package.seek(0)
        uploaded = self.client.post("/api/admin/course-packages", headers=self.admin_headers(), data={
            "classId": "class-one", "title": "第五课", "subtitle": "列表练习", "sequence": "5",
            "syncAssessment": "true", "category": "循环", "file": (package, "lesson-five.zip", "application/zip"),
        }, content_type="multipart/form-data")
        self.assertEqual(uploaded.status_code, 200, uploaded.json)
        self.assertIsNotNone(uploaded.json["assessmentResourceId"])
        page = self.client.get(uploaded.json["path"])
        self.assertEqual(page.status_code, 200)
        self.assertIn(b"\xe7\xac\xac\xe4\xba\x94\xe8\xaf\xbe", page.data)
        page.close()
        state = self.client.get("/api/admin/feature-state?class_id=class-one", headers=self.admin_headers()).json
        self.assertTrue(any(item["title"] == "第五课" and item["published"] for item in state["courses"]))
        self.assertTrue(any(item["title"] == "第五课课后测评" and item["published"] for item in state["assessments"]))
        uploaded_course = next(item for item in state["resourceLibrary"] if item["id"] == uploaded.json["courseResourceId"])
        self.assertEqual(uploaded_course["category"], "循环")
        uploaded_assessment = next(item for item in state["resourceLibrary"] if item["id"] == uploaded.json["assessmentResourceId"])
        self.assertEqual(uploaded_assessment["category"], "循环")
        changed = self.client.put(f"/api/admin/class-resources/{uploaded_assessment['id']}", headers=self.admin_headers(), json={
            "classId": "class-one", "assigned": True, "enabled": True, "category": "课后测评",
        })
        self.assertEqual(changed.status_code, 200)
        refreshed = self.client.get("/api/admin/feature-state?class_id=class-one", headers=self.admin_headers()).json
        self.assertEqual(next(item for item in refreshed["resourceLibrary"] if item["id"] == uploaded_assessment["id"])["category"], "课后测评")

    def test_attendance_start_auto_checkin_manual_leave_and_export(self):
        from openpyxl import load_workbook
        today = __import__("datetime").datetime.now().astimezone().date().isoformat()
        started = self.client.post("/api/admin/attendance/sessions", headers=self.admin_headers(), json={
            "classId": "class-one", "className": "Python班", "termId": "term-one", "termName": "暑假班",
            "date": today, "courseId": "class-one:lesson-1", "courseTitle": "第一课",
            "roster": [{"studentName": "林小满", "accountId": "demo-student-01"}],
        })
        self.assertEqual(started.status_code, 201, started.json)
        session_id = started.json["session"]["id"]
        self.client.post("/api/student/login", json={"username": "student01", "password": "123456"})
        touched = self.client.post("/api/student/attendance/touch", json={"classId": "class-one"})
        self.assertTrue(touched.json["checkedIn"])
        state = self.client.get("/api/admin/attendance?class_id=class-one&term_id=term-one", headers=self.admin_headers()).json
        record = next(item for item in state["sessions"][0]["records"] if item["accountId"] == "demo-student-01")
        self.assertEqual(record["status"], "present")
        changed = self.client.patch(
            f"/api/admin/attendance/sessions/{session_id}/records/{record['studentKey']}",
            headers=self.admin_headers(), json={"status": "leave"},
        )
        self.assertEqual(changed.json["session"]["counts"]["leave"], 1)
        exported = self.client.get("/api/admin/attendance/export?class_id=class-one&term_id=term-one", headers=self.admin_headers())
        self.assertEqual(exported.status_code, 200)
        self.assertTrue(exported.data.startswith(b"PK"))
        workbook = load_workbook(io.BytesIO(exported.data), read_only=True)
        self.assertEqual(workbook.active["A1"].value, "学生姓名")
        self.assertIn("第一课", workbook.active["B1"].value)
        workbook.close()
        exported.close()

        db = sqlite3.connect(self.db_path)
        db.execute("update student_accounts set student_id='profile-one',class_ids='[\"class-two\"]' where id='demo-student-01'")
        db.commit(); db.close()
        self.client.post("/api/parent/login", json={"username": "astudent01", "password": "123456"})
        parent_view = self.client.get("/api/parent/attendance?term_id=term-one")
        self.assertEqual(parent_view.status_code, 200)
        self.assertEqual(parent_view.json["groups"][0]["classId"], "class-one")
        self.assertEqual(parent_view.json["groups"][0]["leave"], 1)

    def test_student_xlsx_preview_generates_initials_and_collision_numbers(self):
        from openpyxl import Workbook, load_workbook
        template = self.client.get("/api/admin/student-import/template", headers=self.admin_headers())
        template_book = load_workbook(io.BytesIO(template.data), read_only=True)
        self.assertEqual(template_book.active.max_row, 1)
        self.assertEqual(template_book.active["A1"].value, "学生姓名")
        template_book.close(); template.close()
        workbook = Workbook(); sheet = workbook.active
        sheet.append(["学生姓名"]); sheet.append(["杨金鹏"]); sheet.append(["杨金鹏"]); sheet.append([None])
        payload = io.BytesIO(); workbook.save(payload); payload.seek(0)
        response = self.client.post("/api/admin/student-import/preview", headers=self.admin_headers(), data={
            "file": (payload, "students.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        }, content_type="multipart/form-data")
        self.assertEqual(response.status_code, 200, response.json)
        self.assertEqual([item["username"] for item in response.json["rows"]], ["yjp", "yjp1"])
        self.assertEqual(response.json["rows"][0]["parentUsername"], "ayjp")


if __name__ == "__main__":
    unittest.main()
